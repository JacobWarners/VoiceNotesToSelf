from datetime import datetime
from openpyxl import load_workbook
now = datetime.now()

def putDataToSpreadsheet(audioRecording):
    myFileName=r'test.xlsx'
    #load the workbook, and put the sheet into a variable
    wb = load_workbook(filename=myFileName)
    ws = wb['Sheet1']

#max_row is a sheet function that gets the last row in a sheet.
    newRowLocation = ws.max_row +1

    #write to the cell you want, specifying row and column, and value :-)
    ws.cell(column=1,row=newRowLocation, value=str(audioRecording))
    ws.cell(column=2,row=newRowLocation, value=now.strftime("%d/%m/%Y %H:%M:%S"))
    ws.cell(column=3,row=newRowLocation, value=str('firstword'))

    wb.save(filename=myFileName)
    wb.close()

putDataToSpreadsheet(audioRecording='test')

# def putDataToSpreadsheet(audioRecording):
#     d = {'col1': [str(audioRecording)], 'col2': [)],'col3': [str('firstword')]}
#     df = pd.DataFrame(data=d)
#     print(df)
#     df.to_excel('test.xlsx')
# putDataToSpreadsheet(audioRecording='test')
